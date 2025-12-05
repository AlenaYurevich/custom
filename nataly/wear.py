import os
from openpyxl import load_workbook
from django.core.cache import cache


EXCEL_FILE = 'wear.xlsx'
CACHE_KEY = 'excel_data'
CACHE_TIMEOUT = 3600  # 1 час


def get_excel_tables():
    """
    Загружает данные из всех листов с кэшированием
    """
    tables = cache.get(CACHE_KEY)
    if tables:
        return tables
    file_path = os.path.join(os.path.dirname(__file__), 'files', EXCEL_FILE)
    workbook = load_workbook(file_path, data_only=True)

    tables = {
        'main': workbook['Table 1'],
        'elements': workbook['Table 2'],
    }

    cache.set(CACHE_KEY, tables, CACHE_TIMEOUT)
    return tables


def fabric_matches(input_fabric, table_fabric):
    """Проверяет совпадение ткани с учетом диапазонов"""
    input_fabric = str(input_fabric).strip()
    table_fabric = str(table_fabric).strip()

    if input_fabric == table_fabric:
        return True

    if '-' in table_fabric:
        try:
            start, end = map(int, table_fabric.split('-'))
            input_num = int(input_fabric)
            return start <= input_num <= end
        except (ValueError, IndexError):
            return False

    return False


def get_value(product_type, fabric, sheet):
    """Получает значение из 5-й колонки указанного листа"""
    current_group = None
    fabric_str = str(fabric).strip()
    use_coefficient = fabric_str == '0'
    search_fabric = '1' if use_coefficient else fabric_str

    for row in range(3, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=1).value

        if cell_value:
            current_group = str(cell_value).strip()

        if not current_group:
            continue

        if current_group == str(product_type).strip():
            fabric_value = sheet.cell(row=row, column=2).value

            if fabric_value and fabric_matches(search_fabric, str(fabric_value).strip()):
                value_cell = sheet.cell(row=row, column=5).value

                if value_cell:
                    try:
                        value = float(str(value_cell).replace(',', '.'))
                        return value * 1.2 if use_coefficient else value
                    except (ValueError, TypeError):
                        return 0

    return 0


def get_value_from_table1(product_type, fabric):
    """Получает значение из Table 1"""
    tables = get_excel_tables()
    return get_value(product_type, fabric, tables['main'])


def get_value_for_element(element, fabric):
    """Получает значение из Table 2"""
    tables = get_excel_tables()
    return get_value(element, fabric, tables['elements'])


def read_types(wearsheet):
    """Читает уникальные значения из первого столбца"""
    choices = []

    for i in range(3, wearsheet.max_row + 1):
        cell = wearsheet.cell(row=i, column=1).value
        if cell:
            value = str(cell).strip()
            if value and value not in choices:
                choices.append(value)

    return tuple((choice, choice) for choice in choices)


def get_choices_type():
    """Получает типы продуктов из Table 1"""
    excel_data = get_excel_tables()
    return read_types(excel_data['main']) if excel_data else ()


def get_elements_type():
    """Получает элементы из Table 2"""
    excel_data = get_excel_tables()
    return read_types(excel_data['elements']) if excel_data else ()


Choices_type = get_choices_type()
Elements_type = get_elements_type()
